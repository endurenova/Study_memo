import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

wb = Workbook(write_only=True)  # 엑셀 파일을 워크북이라 하고
# 파일에서 데이터를 읽어오지 않고, 데이터를 저장하기만 하는 경우에 쓰는 write_only 모드는 데이터를 엑셀 파일에 쓰는데 최적화된 모드
ws = wb.create_sheet('TV Ratings')  # 시트 생성 및 네이밍
ws.append(['순위', '채널', '프로그램', '시청률'])  # 행 A, 행 B, 행 C, 행 D

response = requests.get("https://workey.codeit.kr/ratings/index", verify=False)
rating_page = response.text

soup = BeautifulSoup(rating_page, 'html.parser')
for tr_tag in soup.select('tr')[1:]:  # 첫 번째 인덱스, 즉 첫 번째 행은 헤더라서 1번 부터
    td_tags = tr_tag.select('td')
    row = [
        td_tags[0].get_text(),  # 순위
        td_tags[1].get_text(),  # 채널
        td_tags[2].get_text(),  # 프로그램
        td_tags[3].get_text(),  # 시청률
    ]
    ws.append(row)

wb.save('시청률_2010년1월1주차.xlsx')  # 워크북 저장

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

wb = Workbook(write_only=True)  # 엑셀 파일을 워크북이라 하고
# 파일에서 데이터를 읽어오지 않고, 데이터를 저장하기만 하는 경우에 쓰는 write_only 모드는 데이터를 엑셀 파일에 쓰는데 최적화된 모드

for year in(2010, 2019):
    ws = wb.create_sheet(f'{year}년')  # 시트 생성 및 네이밍
    ws.append(['기간', '순위', '채널', '프로그램', '시청률'])  # 행 A, 행 B, 행 C, 행 D, 행 E

    for month in range(1, 13):
        for weekIndex in range(0, 5):
            response = requests.get(f"https://workey.codeit.kr/ratings/index?year={year}&month={month}&weekIndex={weekIndex}", verify=False)
            rating_page = response.text
            soup = BeautifulSoup(rating_page, 'html.parser')

            for tr_tag in soup.select('tr')[1:]:  # 첫 번째 인덱스, 즉 첫 번째 행은 헤더라서 1번 부터
                td_tags = tr_tag.select('td')
                period = f"{year}년 {month}월 {weekIndex + 1}주차"
                row = [
                    period,
                    td_tags[0].get_text(),  # 순위
                    td_tags[1].get_text(),  # 채널
                    td_tags[2].get_text(),  # 프로그램
                    td_tags[3].get_text(),  # 시청률
                ]
                ws.append(row)

wb.save('시청률.xlsx')  # 워크북 저장


"""CSV (Comma Seprated Values) - 쉼표로 구분되어 있는 파일 형식"""
# name, height, weight
# James, 175, 80
# Alice, 160, 54
# Matthew, 180, 78
# Alex, 173, 72
# Rachel, 164, 58
# 메모장으로 열면 위 같은 형식이나 Excel, Numbers, pandas 같은 데이터 분석 툴로 열면 표로 인식


import csv
import requests
from bs4 import BeautifulSoup

# CSV 파일 생성
csv_file = open('시청률_2010년1월1주차.csv', 'w')
csv_writer = csv.writer(csv_file)

# 헤더 행 추가
csv_writer.writerow(['순위', '채널', '프로그램', '시청률'])

response = requests.get("https://workey.codeit.kr/ratings/index")
rating_page = response.text

soup = BeautifulSoup(rating_page, 'html.parser')

for tr_tag in soup.select('tr')[1:]:
    td_tags = tr_tag.select('td')
    row = [
        td_tags[0].get_text(),
        td_tags[1].get_text(),
        td_tags[2].get_text(),
        td_tags[3].get_text()
    ]
    # 데이터 행 추가
    csv_writer.writerow(row)

# CSV 파일 닫기
csv_file.close()